from bs4 import BeautifulSoup
import requests
import openpyxl as xl

book = xl.Workbook()
sheet = book.active

f = open('D:\\SortComb_62.txt', 'r', encoding='utf-8')

x, y = 2, 2
data = []

for line in f:
    initials = "".join([i for i in list(line.strip()) if not i.isnumeric()])
    print(initials)
    url = f"https://en.wikipedia.org/w/index.php?search={initials}&title=Special%3ASearch&ns0=1"
    sheet.cell(row=x, column=1).value = line.strip()
    scrap = requests.get(url)
    soup = BeautifulSoup(scrap.text, 'html.parser')
    searches = soup.findAll('li', class_="mw-search-result mw-search-result-ns-0")
    if searches:
        initials = str(searches[0])
        initials = initials[initials.find('wiki/') + 5:]
        string = initials[:initials.find('"')]
        new_url = f"https://en.wikipedia.org/w/index.php?search={initials}&title=Special%3ASearch&ns0=1"
        scrap = requests.get(new_url)
        soup = BeautifulSoup(scrap.text, 'html.parser')
    infoBoxData = soup.findAll('td', class_="infobox-data")
    infoBoxLabel = soup.findAll('th', class_="infobox-label")
    for i in range(len(infoBoxLabel)):
        if infoBoxLabel[i].text not in data:
            sheet.cell(row=1, column=y).value = infoBoxLabel[i].text
            sheet.cell(row=x, column=y).value = infoBoxData[i].text
            data.append(infoBoxLabel[i].text)
            y += 1
        else:
            sheet.cell(row=x, column=(data.index(infoBoxLabel[i].text) + 2)).value = infoBoxData[i].text
    x += 1
f.close()

book.save("Matematiki_62.xlsx")
book.close()